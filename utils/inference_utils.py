import os
import re
import time
import sklearn
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sn
from openpyxl import load_workbook

def display_cm(weight_path, measure, data, data_name, ground_truth, predictions):
    rows_lab = []
    rows = []
    for el in ground_truth:
        if el not in rows:
            rows.append(el)
            rows_lab.append(list(data.conversion.keys())[el])
    rows = sorted(rows)
    rows_lab = sorted(rows_lab)
    
    # Confusion matrix based on top 1 accuracy
    columns = []
    columns_lab = []
    for el in predictions[0]:
        if el not in columns:
            columns.append(el)
            columns_lab.append(list(data.conversion.keys())[el])
    columns = sorted(columns)
    columns_lab=sorted(columns_lab)
        
    cm = sklearn.metrics.confusion_matrix(ground_truth, predictions, labels=range(len(os.listdir(data.root)))) # classes predites = colonnes
    # ! only working cause the dic is sorted and sklearn is creating cm by sorting the labels
    df_cm = pd.DataFrame(cm[np.ix_(rows, columns)], index=rows_lab, columns=columns_lab)
    plt.figure(figsize = (10,7))
    sn.heatmap(df_cm, annot=True, xticklabels=True, yticklabels=True)
    #plt.show()
    # save the confusion matrix
    fold_path = weight_path[0:weight_path.rfind("/")]
    plt.savefig(fold_path + '/' + data_name+ '_confusion_matrix_top1_'+measure+ '.png')
    # Confusion matrix based on maj_class accuracy:
    columns = []
    columns_lab = []
    for el in predictions[1]:
        if el not in columns:
            columns.append(el)
            columns_lab.append(list(data.conversion.keys())[el])
    columns = sorted(columns)
    columns_lab = sorted(columns_lab)
    cm = sklearn.metrics.confusion_matrix(ground_truth, predictions[1], labels=range(len(os.listdir(data.root)))) # classes predites = colonnes)
    # ! only working cause the dic is sorted and sklearn is creating cm by sorting the labels
    df_cm = pd.DataFrame(cm[np.ix_(rows, columns)], index=rows_lab, columns=columns_lab)
    plt.figure(figsize = (10,7))
    sn.heatmap(df_cm, annot=True, xticklabels=True, yticklabels=True)
    #plt.show()
    # save the confusion matrix
    plt.savefig(fold_path + '/uliege_confusion_matrix_maj_'+measure+'.png')

def display_precision_recall(weight, measure, ground_truth, predictions):
    from sklearn.metrics import average_precision_score, precision_recall_curve
    from sklearn.preprocessing import label_binarize
    from sklearn.metrics import PrecisionRecallDisplay

    classes = np.unique(ground_truth)
    Y_test = label_binarize(ground_truth, classes=classes)
    y_score = label_binarize(predictions, classes=classes)
    # For each class
    precision = dict()
    recall = dict()
    average_precision = dict()
    for i in range(67):
        precision[i], recall[i], _ = precision_recall_curve(Y_test[:, i], y_score[:, i])
        average_precision[i] = average_precision_score(Y_test[:, i], y_score[:, i])

    # A "micro-average": quantifying score on all classes jointly
    precision["micro"], recall["micro"], _ = precision_recall_curve(
        Y_test.ravel(), y_score.ravel()
    )
    average_precision["micro"] = average_precision_score(Y_test, y_score, average="micro")

    from collections import Counter

    display = PrecisionRecallDisplay(
        recall=recall["micro"],
        precision=precision["micro"],
        average_precision=average_precision["micro"],
        prevalence_pos_label=Counter(Y_test.ravel())[1] / Y_test.size,
    )
    display.plot(plot_chance_level=True)
    _ = display.ax_.set_title("Micro-averaged over all classes")
    fold_path = weight[0:weight.rfind("/")]
    plt.savefig(fold_path + '/uliege_prec_recall_curve_'+measure+'.png')
    
def display_prec_im(weight, props, data, measure):
    plt.figure()    
    props = props / data.__len__()
    plt.plot(props)
    plt.xlabel('Number of images')
    plt.ylabel('Proportion of correct images')
    fold_path = weight[0:weight.rfind("/")]
    plt.savefig(fold_path + '/uliege_prec_im_'+measure+'.png')

def safe_stat(x):
    """Return either scalar, mean ± std, or list of mean ± std if shape[0] == 3."""
    try:
        arr = np.array(x)
        # Stat case for accuracies: (3, nb_exp)
        if arr.ndim == 2 and arr.shape[0] == 3:
            stats = []
            for row in arr:
                stats.append(f"{np.mean(row):.4f} ± {np.std(row):.4f}")
            return stats  

        # Stat case for times or other (nb_exp,) (nb_exp > 3)
        elif arr.ndim == 1 and arr.size > 3:
            return f"{np.mean(arr):.4f} ± {np.std(arr):.4f}"

        # Scalar case
        else:
            return f"{float(arr):.4f}"

    except Exception:
        return str(x)

def write_full_results_txt(filename, data_name, data_path, model, measure, accuracies, 
                  t_tot, t_model, t_search, t_transfer, class_name=None, project_name=None, nb_exp=None):
    results = {
        "top_1_acc": safe_stat(accuracies[0]),
        "top_5_acc": safe_stat(accuracies[1]),
        "maj_acc": safe_stat(accuracies[2]),
        "t_tot": safe_stat(t_tot),
        "t_model": safe_stat(t_model),
        "t_search": safe_stat(t_search),
        "t_transfer": safe_stat(t_transfer),
    }

    with open(filename, 'a') as f:  # append
        f.write("------------------------------------------------------------------------\n")
        f.write("Results of the inference:\n")
        f.write(f"Model: {model.model_name}\n")
        f.write(f"Measure: {measure}\n")
        f.write("Date and time: " + time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()) + "\n")
        if class_name is not None:
            f.write(f"Class: {class_name}\n")
        if project_name is not None:
            f.write(f"Project: {project_name}\n")
        if nb_exp is not None:
            f.write(f"Number of experiments: {nb_exp}\n")
        f.write(f"Data: {data_name} found in Data path: {data_path}\n\n")
        for key, value in results.items():
            f.write(f"{key}: {value}\n")

def write_full_results_xlsx(filename, data_name, data_path, model, measure, accuracies, 
                            t_tot, t_model, t_search, t_transfer, class_name=None, project_name=None, nb_exp=None):
    # Metadata row as a DataFrame
    meta = {
        "model_name": model.model_name,
        "date": time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()),
        "data_name": data_name,
        "data_path": data_path,
        "measure": measure,
    }
    if class_name is not None:
        meta["class_name"] = class_name
    if project_name is not None:
        meta["project_name"] = project_name
    if nb_exp is not None:
        meta["nb_exp"] = nb_exp

    meta_df = pd.DataFrame([meta])
    
    if data_name == 'uliege': 
        results = np.zeros(13)
        for i in range(9):
            results[i] = safe_stat(accuracies[i//3][i%3])
        results[9] = safe_stat(t_tot)
        results[10] = safe_stat(t_model)
        results[11] = safe_stat(t_search)
        results[12] = safe_stat(t_transfer)
        df = pd.DataFrame([results], columns=["top_1_acc", "top_5_acc", "maj_acc", "top_1_proj", "top_5_proj", "maj_acc_proj", "top_1_sim", "top_5_sim", "maj_acc_sim", "t_tot", "t_model", "t_search", "t_transfer"])
    else:
        results = np.zeros(7)
        for i in range(3):
            results[i] = safe_stat(accuracies[i])
        results[3] = safe_stat(t_tot)
        results[4] = safe_stat(t_model)
        results[5] = safe_stat(t_search)
        results[6] = safe_stat(t_transfer)
        df = pd.DataFrame([results], columns=["top_1_acc", "top_5_acc", "maj_acc", "t_tot", "t_model", "t_search", "t_transfer"])

    # Concatenate metadata row + results
    final_df = pd.concat([meta_df, df], ignore_index=True, axis=1)
    
    if not os.path.exists(filename):
        # Case 1: file does not exist
        with pd.ExcelWriter(filename, engine="openpyxl") as writer:
            final_df.to_excel(writer, sheet_name=data_name, index=False, header=True)
    else:
        # Case 2: append to existing file
        book = load_workbook(filename)
        if data_name in book.sheetnames:
            startrow = book[data_name].max_row
        else:
            startrow = 0  

        with pd.ExcelWriter(filename, mode="a", engine="openpyxl", if_sheet_exists="overlay") as writer:
            final_df.to_excel(writer, sheet_name=data_name,
                              index=False, startrow=startrow, header=(startrow==0))


def write_class_results_xlsx(filename, data_name, data_path, model, results, classes):
    # Metadata row as a DataFrame
    meta = {
        "model_name": model.model_name,
        "date": time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()),
        "data_name": data_name,
        "data_path": data_path,
    }
    meta_df = pd.DataFrame([meta])
    
    if data_name == 'uliege': 
        df = pd.DataFrame(results, columns=["top_1_acc", "top_5_acc", "maj_acc_class", "top_1_proj", "top_5_proj", "maj_acc_proj", "top_1_sim", "top_5_sim", "maj_acc_sim", "t_tot", "t_model", "t_search", "t_transfer"])
    else:
        df = pd.DataFrame(results, columns=["top_1_acc", "top_5_acc", "maj_acc", "t_tot", "t_model", "t_search", "t_transfer"])
    df.index = classes

    # Concatenate metadata row + results
    final_df = pd.concat([meta_df, df], ignore_index=True)
    
    if not os.path.exists(filename):
        # Case 1: file does not exist
        with pd.ExcelWriter(filename, engine="openpyxl") as writer:
            final_df.to_excel(writer, sheet_name="Results_per_class_"+model.model_name, index=False, header=True)
    else:
        # Case 2: append to existing file
        book = load_workbook(filename)
        if "Results_per_class_"+model.model_name in book.sheetnames:
            startrow = book["Results_per_class_"+model.model_name].max_row
        else:
            startrow = 0  

        with pd.ExcelWriter(filename, mode="a", engine="openpyxl", if_sheet_exists="overlay") as writer:
            final_df.to_excel(writer, sheet_name="Results_per_class_"+model.model_name,
                              index=False, startrow=startrow, header=(startrow==0))

def get_weight(model_name, weights_dir):
    version = "model"
    if "dino" in model_name or "ibot" in model_name or "byol" in model_name:
        idx = model_name.find("_")
        version = model_name[idx+1:]
        model_name = model_name[0:idx]
    else:
        numbers = re.findall(r'\d+', model_name)
        if len(numbers) > 0:
            version = "v" + numbers[0]
            model_name = model_name[0:model_name.find(version)-1]
    dir = os.path.join(weights_dir, model_name, version)
    for file in os.listdir(dir):
        if os.path.splitext(file)[0] == "weight":
            weight_path = os.path.join(dir, file)
            return weight_path
    print("No weight found for the model " + model_name + " in the folder " + dir)
    return None